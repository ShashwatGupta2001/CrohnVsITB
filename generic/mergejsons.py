import json
import os
import sys
import re
from openpyxl import Workbook, load_workbook

import re

def preprocess(value):
    if value is None:
        return value
    else:
        value_str = str(value)
        value_str = re.sub(r'[ ,;:\'"]', '', value_str)
        return value_str
    

def merge_jsons(key, files, prefixes):
    merged_data = {}
    
    for file, prefix in zip(files, prefixes):
        with open(file, 'r') as f:
            data = json.load(f)
            print(data.keys())
            if key in data:
                print(11)
                prefixed_data = {f"{prefix}.{k}": preprocess(v) for k, v in data[key].items()}
                merged_data.update(prefixed_data)
    
    return {key: merged_data}

def write_to_xlsx(data, output_file):
    if not data:
        print("No data to write to XLSX.")
        return
    
    directory = list(data.keys())[0]
    merged_data = data[directory]
    row = [directory] + list(merged_data.values())
    file_exists = os.path.isfile(output_file)

    if file_exists:
        workbook = load_workbook(output_file)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            'directory',
            'z_values.min',
            'z_values.max',
            'z_values.mean',
            'z_values.time',
            'z_values.disease',
            'ts.area_values.use_mean_sv',
            'ts.area_values.use_mean_vv',
            'ts.area_values.use_mean_tsv',
            'ts.area_values.use_mean_ratio',
            'ts.area_values.use_minmax_sv',
            'ts.area_values.use_minmax_vv',
            'ts.area_values.use_minmax_tsv',
            'ts.area_values.use_minmax_ratio',
            'ts.area_values.elapsed_time',
            "fr.area_values.use_mean_sv",
            "fr.area_values.use_mean_vv",
            "fr.area_values.use_mean_tsv",
            'fr.area_values.use_mean_ratio',
            "fr.area_values.use_minmax_sv",
            "fr.area_values.use_minmax_vv",
            "fr.area_values.use_minmax_tsv",
            'fr.area_values.use_minmax_ratio',
            "fr.area_values.elapsed_time",
            "dcmhead.Pitch",
            "dcmhead.Acquisition Type",
            "dcmhead.Acquisition Length",
            "dcmhead.Acquisition Duration",
            "dcmhead.Number of Study Related Instances",
            "dcmhead.Manufacturer",
            "dcmhead.Patients Sex",
            "dcmhead.Contrast Agent",
        ]
        sheet.append(headers)
    
    sheet.append(row)
    workbook.save(output_file)

def main():
    key = sys.argv[1]

    json_files = ['z_values.json', 'ts.area_values.json', 'fr.area_values.json', 'dcmhead.json']
    prefixes = ['z_values', 'ts.area_values', 'fr.area_values', 'dcmhead']

    merged_result = merge_jsons(key, json_files, prefixes)

    print(json.dumps(merged_result, indent=4))

    output_file = 'merged_data.xlsx'
    write_to_xlsx(merged_result, output_file)
    
    print(f'Merged data appended to {output_file}')

if __name__ == "__main__":
    main()
